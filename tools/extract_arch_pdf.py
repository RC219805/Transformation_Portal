#!/usr/bin/env python3
"""
extract_arch_pdf.py

Extract + organize data from an architectural PDF (plan set) into a structured
output directory.

It extracts:
  ✅ Document metadata (JSON)
  ✅ Per-page plain text (TXT)
  ✅ Optional per-page layout-aware text lines (JSON: text + bbox + font size)
  ✅ Title block info (heuristics): sheet number, sheet title, project number,
     project address, date, scale (plus raw title block text for auditability)
  ✅ Optional sheet index parsing (sheet number + title pairs)
  ✅ Optional table extraction (PDFPlumber; CSV + JSON; plus tables_index.json)
  ✅ Optional embedded image extraction (PyMuPDF extract_image)
  ✅ Optional page renders (PNG) and OCR fallback (pytesseract)

Recommended run for an architectural plan set:
  python extract_arch_pdf.py "24098.00_750 PICACHO LANE.pdf" \
    --out picacho_extract --jsonl --layout-json --sheet-index --tables --excel

Dependencies (core):
  pip install pymupdf pdfplumber

Optional:
  pip install pytesseract pillow camelot-py[cv] openpyxl

Notes:
- OCR requires the Tesseract binary installed and on PATH.
- Table extraction quality varies widely; many schedules are raster images.
- This script keeps raw extracted text alongside parsed fields to make OCR quirks
  inspectable and fixable.
"""

from __future__ import annotations

import argparse
import csv
import json
import logging
import re
from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence, Tuple

import fitz  # PyMuPDF

# Optional imports (used only when the corresponding flags are enabled)
try:
    import pdfplumber  # type: ignore
except Exception:  # pragma: no cover
    pdfplumber = None  # type: ignore

try:
    import camelot  # type: ignore
except Exception:  # pragma: no cover
    camelot = None  # type: ignore

try:
    import pytesseract  # type: ignore
except Exception:  # pragma: no cover
    pytesseract = None  # type: ignore

try:
    import openpyxl  # type: ignore
    from openpyxl.utils import get_column_letter  # type: ignore
except Exception:  # pragma: no cover
    openpyxl = None  # type: ignore
    get_column_letter = None  # type: ignore


LOGGER = logging.getLogger("extract_arch_pdf")


# -----------------------------
# Data models
# -----------------------------
@dataclass(frozen=True)
class TextLine:
    """A single text line with page coordinates."""

    text: str
    bbox: Tuple[float, float, float, float]  # (x0, y0, x1, y1)
    max_font_size: float


@dataclass(frozen=True)
class TitleBlockInfo:
    region_name: str
    region_bbox: Tuple[float, float, float, float]
    raw_text: str

    sheet_no: Optional[str] = None
    sheet_title: Optional[str] = None
    project_no: Optional[str] = None
    project_address: Optional[str] = None
    date: Optional[str] = None
    scale: Optional[str] = None


# -----------------------------
# Helpers
# -----------------------------
SHEET_ID_PATTERNS: Sequence[re.Pattern[str]] = (
    # A101, L401, MEP1, AD1.02, A1.07, S0.1, S-0.1
    re.compile(r"^[A-Z]{1,4}-?\d{1,4}(?:\.\d{1,3})*$"),
    # T.01 (letters + dot + digits)
    re.compile(r"^[A-Z]{1,4}\.\d{1,3}(?:\.\d{1,3})*$"),
)

TITLE_KEYWORDS = (
    "PLAN",
    "ELEVATION",
    "SECTION",
    "DETAIL",
    "DETAILS",
    "SCHEDULE",
    "NOTES",
    "LEGEND",
    "SPEC",
    "SPECIFICATIONS",
    "DIAGRAM",
    "SITE",
    "FLOOR",
    "ROOF",
    "FOUNDATION",
    "FRAMING",
    "CEILING",
    "RCP",
    "DOOR",
    "WINDOW",
    "FINISH",
    "MATERIAL",
    "STRUCTURAL",
    "ELECTRICAL",
    "PLUMBING",
    "MECHANICAL",
    "HVAC",
    "FIRE",
    "LANDSCAPE",
    "CIVIL",
)

KEYWORDS_FOR_TITLEBLOCK = (
    "SHEET",
    "PROJECT",
    "ADDRESS",
    "SCALE",
    "DRAWN",
    "CHECKED",
    "DATE",
    "OWNER",
    "ARCHITECT",
    "ENGINEER",
    "CONSULTANT",
    "REVISION",
    "REVISIONS",
)

ADDRESS_SUFFIXES = (
    "LANE",
    "LN",
    "ROAD",
    "RD",
    "STREET",
    "ST",
    "AVENUE",
    "AVE",
    "DRIVE",
    "DR",
    "COURT",
    "CT",
    "PLACE",
    "PL",
    "BOULEVARD",
    "BLVD",
    "WAY",
    "CIRCLE",
    "CIR",
)


def ensure_dir(path: Path) -> None:
    path.mkdir(parents=True, exist_ok=True)


def write_text(path: Path, content: str) -> None:
    path.write_text(content, encoding="utf-8", errors="replace")


def write_json(path: Path, obj: Any) -> None:
    path.write_text(json.dumps(obj, indent=2, ensure_ascii=False), encoding="utf-8")


def normalize_text(s: str) -> str:
    """Normalize whitespace + strip common OCR artifacts that break parsing."""
    s = s.replace("\u00ad", "")  # soft hyphen
    s = s.replace("\u200b", "")  # zero-width space
    return re.sub(r"\s+", " ", s).strip()


def normalize_sheet_id(raw: str) -> str:
    """Normalize OCR variants like 'S1 .1' -> 'S1.1', 'A1.0 7' -> 'A1.07'."""
    cleaned = re.sub(r"[^A-Za-z0-9\.\-]", "", raw.upper())
    cleaned = re.sub(r"\.{2,}", ".", cleaned)
    cleaned = cleaned.strip(".-")
    return cleaned


def is_sheet_id(candidate: str) -> bool:
    c = normalize_sheet_id(candidate)
    if not c or not re.search(r"\d", c):
        return False
    return any(p.match(c) for p in SHEET_ID_PATTERNS)


def letters_upper_ratio(s: str) -> float:
    letters = [ch for ch in s if ch.isalpha()]
    if not letters:
        return 0.0
    upp = sum(1 for ch in letters if ch.isupper())
    return upp / max(len(letters), 1)


def title_likeness_score(text: str, font_size: float) -> float:
    """
    Score a candidate sheet title line.

    Prefer short-ish, title-like phrases (often ALL CAPS) and avoid long sentences
    from general notes that happen to be near the title block.
    """
    t = text.strip()
    if not t:
        return -1e9

    score = 0.0

    # Length heuristics
    L = len(t)
    if 5 <= L <= 45:
        score += 12.0
    elif 46 <= L <= 70:
        score += 4.0
    else:
        score -= 10.0

    # Uppercase ratio
    score += 6.0 * letters_upper_ratio(t)

    # Keyword boosts
    upper = t.upper()
    for kw in TITLE_KEYWORDS:
        if kw in upper:
            score += 8.0

    # Penalize sentence-like patterns
    if ":" in t:
        score -= 6.0
    if ";" in t:
        score -= 4.0
    if t.endswith("."):
        score -= 6.0
    if "," in t:
        score -= 2.0
    if re.search(r"\b\d{3,}\b", t):
        score -= 3.0

    # Mild font-size tie-breaker
    score += min(font_size, 20.0) * 0.3

    return score


def score_titleblock_text(text: str) -> int:
    t = text.upper()
    score = 0
    for kw in KEYWORDS_FOR_TITLEBLOCK:
        if kw in t:
            score += 3
    tokens = re.findall(r"\b[A-Z]{1,4}[A-Z0-9\.\-]{0,10}\b", t)
    score += sum(1 for tok in tokens if is_sheet_id(tok)) * 6
    # Cap contribution from sheer amount of text (prevents giant regions from winning)
    score += min(len(t) // 250, 8)
    return score


def extract_lines(page: fitz.Page, clip: Optional[fitz.Rect] = None) -> List[TextLine]:
    d = page.get_text("dict", clip=clip)
    out: List[TextLine] = []
    for block in d.get("blocks", []):
        if block.get("type") != 0:
            continue
        for line in block.get("lines", []):
            spans = line.get("spans", [])
            if not spans:
                continue
            text = "".join(span.get("text", "") for span in spans)
            text = normalize_text(text)
            if not text:
                continue
            bbox = tuple(float(x) for x in line.get("bbox", (0, 0, 0, 0)))  # type: ignore
            max_size = float(max(span.get("size", 0.0) for span in spans))
            out.append(TextLine(text=text, bbox=bbox, max_font_size=max_size))
    out.sort(key=lambda l: (l.bbox[1], l.bbox[0]))
    return out


def candidate_titleblock_regions(page: fitz.Page) -> List[Tuple[str, fitz.Rect]]:
    """
    Plausible title block regions. Architectural sets vary:
    - bottom-right title blocks are common
    - some are vertical on the right edge
    """
    w, h = float(page.rect.width), float(page.rect.height)

    bottom_right = fitz.Rect(w * 0.55, h * 0.72, w, h)
    tight_bottom_right = fitz.Rect(w * 0.70, h * 0.78, w, h)
    lower_right = fitz.Rect(w * 0.62, h * 0.60, w, h)
    right_strip = fitz.Rect(w * 0.78, 0, w, h)
    bottom_strip_right = fitz.Rect(w * 0.35, h * 0.78, w, h)
    bottom_strip_full = fitz.Rect(0, h * 0.82, w, h)  # last resort

    return [
        ("tight_bottom_right", tight_bottom_right),
        ("bottom_right", bottom_right),
        ("lower_right", lower_right),
        ("bottom_strip_right", bottom_strip_right),
        ("right_strip", right_strip),
        ("bottom_strip_full", bottom_strip_full),
    ]


def pick_titleblock_region(page: fitz.Page) -> Tuple[str, fitz.Rect, str, List[TextLine]]:
    """
    Pick the most likely title block region based on keyword/sheet-id scoring,
    with a penalty for very large regions (to avoid grabbing general notes).
    """
    w, h = float(page.rect.width), float(page.rect.height)
    page_area = w * h

    best: Optional[Tuple[float, float, int, str, fitz.Rect, str, List[TextLine]]] = None
    for name, rect in candidate_titleblock_regions(page):
        raw_text = normalize_text(page.get_text("text", clip=rect))
        lines = extract_lines(page, clip=rect)
        base = score_titleblock_text(raw_text)
        area_frac = (float(rect.width) * float(rect.height)) / max(page_area, 1.0)

        final = float(base) - (area_frac * 150.0)

        # Tie-breakers: higher final, then smaller area, then higher base
        # Note: best is checked for None in assertion below (pylint: disable=unsubscriptable-object)
        if best is None or (final, -area_frac, base) > (best[0], -best[1], best[2]):  # pylint: disable=unsubscriptable-object
            best = (final, area_frac, base, name, rect, raw_text, lines)

    assert best is not None
    _, _, _, name, rect, raw_text, lines = best
    return name, rect, raw_text, lines


def parse_sheet_no_and_title(lines: List[TextLine]) -> Tuple[Optional[str], Optional[str]]:
    # Find sheet ID line by match + max font size
    sheet_line: Optional[TextLine] = None
    for ln in lines:
        if is_sheet_id(ln.text):
            if sheet_line is None or ln.max_font_size > sheet_line.max_font_size:
                sheet_line = ln

    if sheet_line is None:
        return None, None

    sheet_no = normalize_sheet_id(sheet_line.text)

    # Candidate titles near the sheet number line (vertical + horizontal proximity)
    anchor_x = sheet_line.bbox[0]
    anchor_y = sheet_line.bbox[1]

    candidates: List[Tuple[float, TextLine]] = []
    for ln in lines:
        if ln is sheet_line:
            continue
        if is_sheet_id(ln.text):
            continue
        if len(ln.text) < 4:
            continue
        if not re.search(r"[A-Za-z]{3,}", ln.text):
            continue

        # Vertical window near sheet no
        if ln.bbox[1] < anchor_y - 25 or ln.bbox[1] > anchor_y + 170:
            continue

        # Horizontal proximity
        if abs(ln.bbox[0] - anchor_x) > 520:
            continue

        score = title_likeness_score(ln.text, ln.max_font_size)
        candidates.append((score, ln))

    if not candidates:
        return sheet_no, None

    candidates.sort(key=lambda x: x[0], reverse=True)
    best_score, best_line = candidates[0]
    if best_score < 0:
        return sheet_no, None
    return sheet_no, best_line.text


def parse_project_number(lines: List[TextLine]) -> Optional[str]:
    """
    Extract project number from title block lines (requires a PROJECT label to
    reduce false positives).
    """
    for ln in lines:
        if re.search(r"\bPROJECT\b", ln.text, re.IGNORECASE) and "ADDRESS" not in ln.text.upper():
            m = re.search(r"([0-9]{3,}\s*[.\s]\s*[0-9]{2,})", ln.text)
            if not m:
                continue
            raw = m.group(1)
            raw = re.sub(r"\s+", ".", raw.strip())
            raw = re.sub(r"\.{2,}", ".", raw)
            return raw
    return None


def parse_address(lines: List[TextLine]) -> Optional[str]:
    """
    Extract project address from title block lines using label-first parsing.
    """
    # 1) Label-driven: "PROJECT ADDRESS" -> next line(s)
    for i, ln in enumerate(lines):
        if "ADDRESS" in ln.text.upper():
            for j in range(i + 1, min(i + 6, len(lines))):
                cand = lines[j].text.upper()
                if re.search(r"\b\d{1,6}\b", cand) and any(sfx in cand for sfx in ADDRESS_SUFFIXES):
                    return normalize_text(lines[j].text)

    # 2) Pattern fallback
    suffix_group = "|".join(re.escape(s) for s in ADDRESS_SUFFIXES)
    addr_re = re.compile(rf"\b(\d{{1,6}}\s+[A-Z0-9][A-Z0-9\s\-]{{2,60}}\s+(?:{suffix_group}))\b", re.IGNORECASE)
    for ln in lines:
        if "PROJECT" in ln.text.upper():
            continue
        m = addr_re.search(ln.text)
        if m:
            return normalize_text(m.group(1))
    return None


def parse_date(lines: List[TextLine], raw_text: str) -> Optional[str]:
    # Prefer label-based extraction
    for ln in lines:
        if "DATE" in ln.text.upper():
            m = re.search(r"\b(\d{1,2}/\d{1,2}/\d{2,4})\b", ln.text)
            if m:
                return m.group(1)
            m = re.search(r"\b(20\d{2}-\d{1,2}-\d{1,2})\b", ln.text)
            if m:
                return m.group(1)

    # Fallback search anywhere in raw_text
    m = re.search(r"\b(\d{1,2}/\d{1,2}/\d{2,4})\b", raw_text)
    if m:
        return m.group(1)
    m = re.search(r"\b(20\d{2}-\d{1,2}-\d{1,2})\b", raw_text)
    if m:
        return m.group(1)
    return None


def parse_scale(lines: List[TextLine]) -> Optional[str]:
    for i, ln in enumerate(lines):
        if "SCALE" not in ln.text.upper():
            continue

        m = re.search(r"SCALE[:\s]+(.+)$", ln.text, re.IGNORECASE)
        cand = normalize_text(m.group(1)) if m else ""

        if not cand and i + 1 < len(lines):
            cand = normalize_text(lines[i + 1].text)

        if not cand:
            continue

        upper = cand.upper()
        looks_like_scale = (
            "AS NOTED" in upper
            or re.search(r"\d+\s*/\s*\d+", cand) is not None
            or ('"' in cand and "'" in cand)
            or ('"' in cand and "=" in cand)
        )
        if not looks_like_scale:
            continue

        return cand[:80]

    return None


def title_block_info(page: fitz.Page) -> TitleBlockInfo:
    region_name, rect, raw_text, lines = pick_titleblock_region(page)

    sheet_no, sheet_title = parse_sheet_no_and_title(lines)
    project_no = parse_project_number(lines)
    address = parse_address(lines)
    date = parse_date(lines, raw_text)
    scale = parse_scale(lines)

    return TitleBlockInfo(
        region_name=region_name,
        region_bbox=(float(rect.x0), float(rect.y0), float(rect.x1), float(rect.y1)),
        raw_text=raw_text,
        sheet_no=sheet_no,
        sheet_title=sheet_title,
        project_no=project_no,
        project_address=address,
        date=date,
        scale=scale,
    )


# -----------------------------
# Sheet index parsing
# -----------------------------
def parse_sheet_index_from_page_lines(lines: List[TextLine]) -> List[Tuple[str, str]]:
    """
    Parse a sheet index/list page: returns [(sheet_no, sheet_title), ...].
    """
    rows: List[List[TextLine]] = []
    tol = 4.0  # points

    for ln in sorted(lines, key=lambda l: (l.bbox[1], l.bbox[0])):
        cy = (ln.bbox[1] + ln.bbox[3]) / 2.0
        placed = False
        for row in rows:
            row_cy = (row[0].bbox[1] + row[0].bbox[3]) / 2.0
            if abs(cy - row_cy) <= tol:
                row.append(ln)
                placed = True
                break
        if not placed:
            rows.append([ln])

    pairs: List[Tuple[str, str]] = []
    for r_i, row in enumerate(rows):
        row_sorted = sorted(row, key=lambda l: l.bbox[0])

        sheet_items = [ln for ln in row_sorted if is_sheet_id(ln.text)]
        if not sheet_items:
            continue

        sheet_ln = min(sheet_items, key=lambda l: l.bbox[0])
        sheet_no = normalize_sheet_id(sheet_ln.text)

        title_candidates = [
            ln for ln in row_sorted if ln.bbox[0] > sheet_ln.bbox[2] + 5 and re.search(r"[A-Za-z]{3,}", ln.text)
        ]
        title = ""
        if title_candidates:
            title_candidates.sort(key=lambda l: (len(l.text), l.max_font_size), reverse=True)
            title = title_candidates[0].text
        else:
            # Fallback: stacked layout (title below)
            if r_i + 1 < len(rows):
                next_row = sorted(rows[r_i + 1], key=lambda l: l.bbox[0])
                stacked = [
                    ln for ln in next_row if abs(ln.bbox[0] - sheet_ln.bbox[0]) < 60 and re.search(r"[A-Za-z]{3,}", ln.text)
                ]
                if stacked:
                    stacked.sort(key=lambda l: (len(l.text), l.max_font_size), reverse=True)
                    title = stacked[0].text

        if title:
            pairs.append((sheet_no, title))

    seen = set()
    out = []
    for s, t in pairs:
        key = (s, t)
        if key in seen:
            continue
        seen.add(key)
        out.append((s, t))
    return out


# -----------------------------
# Table extraction
# -----------------------------
def extract_tables_pdfplumber(pdf_path: Path, out_dir: Path) -> List[Dict[str, Any]]:
    if pdfplumber is None:
        raise RuntimeError("pdfplumber is not installed; install with: pip install pdfplumber")

    ensure_dir(out_dir)
    index: List[Dict[str, Any]] = []

    with pdfplumber.open(str(pdf_path)) as pdf:
        for page_idx, page in enumerate(pdf.pages):
            try:
                tables = page.extract_tables()
            except Exception as e:  # pragma: no cover
                LOGGER.warning("Table extraction failed on page %s: %s", page_idx + 1, e)
                continue

            for t_idx, table in enumerate(tables, start=1):
                if not table:
                    continue
                if all((cell is None or str(cell).strip() == "") for row in table for cell in row):
                    continue

                csv_path = out_dir / f"page_{page_idx+1:03d}_table_{t_idx:02d}.csv"
                with csv_path.open("w", newline="", encoding="utf-8") as f:
                    writer = csv.writer(f)
                    for row in table:
                        writer.writerow([normalize_text(str(cell)) if cell is not None else "" for cell in row])

                json_path = out_dir / f"page_{page_idx+1:03d}_table_{t_idx:02d}.json"
                write_json(json_path, table)

                index.append(
                    {
                        "engine": "pdfplumber",
                        "page_number": page_idx + 1,
                        "table_number": t_idx,
                        "csv": str(csv_path.name),
                        "json": str(json_path.name),
                        "rows": len(table),
                        "cols": max(len(r) for r in table) if table else 0,
                    }
                )

    return index


def extract_tables_camelot(pdf_path: Path, out_dir: Path) -> List[Dict[str, Any]]:
    if camelot is None:
        raise RuntimeError("camelot is not installed; install with: pip install camelot-py[cv]")

    ensure_dir(out_dir)
    index: List[Dict[str, Any]] = []

    for flavor in ("lattice", "stream"):
        try:
            tables = camelot.read_pdf(str(pdf_path), pages="all", flavor=flavor)
        except Exception as e:  # pragma: no cover
            LOGGER.warning("Camelot (%s) failed: %s", flavor, e)
            continue

        for i, t in enumerate(tables, start=1):
            df = t.df
            if df.empty:
                continue
            csv_path = out_dir / f"camelot_{flavor}_{i:03d}.csv"
            df.to_csv(csv_path, index=False, header=False)
            index.append(
                {
                    "engine": "camelot",
                    "flavor": flavor,
                    "table_number": i,
                    "page": getattr(t, "page", None),
                    "csv": str(csv_path.name),
                    "shape": [int(df.shape[0]), int(df.shape[1])],
                }
            )

    return index


# -----------------------------
# Image extraction / rendering / OCR
# -----------------------------
def extract_embedded_images(doc: fitz.Document, page: fitz.Page, out_dir: Path, page_number: int) -> List[Dict[str, Any]]:
    ensure_dir(out_dir)
    extracted: List[Dict[str, Any]] = []

    images = page.get_images(full=True)
    for idx, img in enumerate(images, start=1):
        xref = img[0]
        try:
            base = doc.extract_image(xref)
        except Exception as e:  # pragma: no cover
            LOGGER.warning("Failed to extract image on page %s (xref=%s): %s", page_number, xref, e)
            continue
        ext = base.get("ext", "bin")
        img_bytes = base.get("image", b"")
        img_path = out_dir / f"page_{page_number:03d}_img_{idx:03d}.{ext}"
        img_path.write_bytes(img_bytes)
        extracted.append(
            {
                "page_number": page_number,
                "image_number": idx,
                "xref": xref,
                "ext": ext,
                "path": str(img_path.name),
                "width": base.get("width"),
                "height": base.get("height"),
                "colorspace": base.get("colorspace"),
            }
        )

    return extracted


def render_page(page: fitz.Page, out_path: Path, dpi: int = 120) -> None:
    zoom = dpi / 72.0
    pix = page.get_pixmap(matrix=fitz.Matrix(zoom, zoom), alpha=False)
    out_path.write_bytes(pix.tobytes("png"))


def ocr_page(page: fitz.Page, dpi: int = 300) -> str:
    if pytesseract is None:
        raise RuntimeError("pytesseract not installed. pip install pytesseract")
    try:
        from PIL import Image  # type: ignore
    except Exception as e:  # pragma: no cover
        raise RuntimeError("Pillow not installed. pip install pillow") from e

    zoom = dpi / 72.0
    pix = page.get_pixmap(matrix=fitz.Matrix(zoom, zoom), alpha=False)
    img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
    return normalize_text(pytesseract.image_to_string(img))


# -----------------------------
# Excel export
# -----------------------------
def write_excel_summary(
    out_xlsx: Path,
    doc_meta: Dict[str, Any],
    sheets: List[Dict[str, Any]],
    sheet_index_pairs: List[Dict[str, Any]],
    tables_index: List[Dict[str, Any]],
) -> None:
    if openpyxl is None:
        raise RuntimeError("openpyxl is not installed; install with: pip install openpyxl")

    wb = openpyxl.Workbook()

    ws_meta = wb.active
    ws_meta.title = "metadata"
    ws_meta.append(["key", "value"])
    for k, v in sorted(doc_meta.items()):
        ws_meta.append([k, json.dumps(v, ensure_ascii=False) if isinstance(v, (dict, list)) else str(v)])

    ws_sheets = wb.create_sheet("sheets")
    if sheets:
        headers = list(sheets[0].keys())
        ws_sheets.append(headers)
        for row in sheets:
            ws_sheets.append([row.get(h, "") for h in headers])

    ws_idx = wb.create_sheet("sheet_index")
    if sheet_index_pairs:
        headers = list(sheet_index_pairs[0].keys())
        ws_idx.append(headers)
        for row in sheet_index_pairs:
            ws_idx.append([row.get(h, "") for h in headers])

    ws_tbl = wb.create_sheet("tables")
    if tables_index:
        headers = list(tables_index[0].keys())
        ws_tbl.append(headers)
        for row in tables_index:
            ws_tbl.append([row.get(h, "") for h in headers])

    for ws in wb.worksheets:
        for col in range(1, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(col)].width = 26  # type: ignore[arg-type]

    wb.save(out_xlsx)


# -----------------------------
# Main extraction
# -----------------------------
def build_arg_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(description="Extract and organize text/tables/images from an architectural PDF.")
    p.add_argument("pdf", type=str, help="Input PDF path")
    p.add_argument("--out", type=str, default="out_extract", help="Output directory")

    p.add_argument(
        "--layout-json", action="store_true", help="Write layout-aware per-page JSON (lines with bboxes/font sizes)."
    )
    p.add_argument("--jsonl", action="store_true", help="Write pages.jsonl (one JSON object per page).")

    p.add_argument("--tables", action="store_true", help="Extract tables via pdfplumber (CSV + JSON).")
    p.add_argument("--camelot", action="store_true", help="Also try Camelot table extraction (optional).")

    p.add_argument("--images", action="store_true", help="Extract embedded images from PDF.")
    p.add_argument("--render-pages", action="store_true", help="Render each page to PNG (can be large).")
    p.add_argument("--render-dpi", type=int, default=120, help="DPI for page renders (when --render-pages).")

    p.add_argument("--ocr-fallback", action="store_true", help="If a page has little/no text, OCR it (slow).")
    p.add_argument(
        "--ocr-min-chars", type=int, default=200, help="If extracted text chars < this, do OCR when --ocr-fallback."
    )
    p.add_argument("--ocr-dpi", type=int, default=300, help="DPI for OCR rasterization.")

    p.add_argument("--sheet-index", action="store_true", help="Attempt to parse sheet index/list pages.")
    p.add_argument("--excel", action="store_true", help="Write an Excel summary workbook.")
    p.add_argument("--max-pages", type=int, default=0, help="Limit number of pages (0 = all).")

    p.add_argument("--log-level", type=str, default="INFO", help="Logging level (DEBUG, INFO, WARNING, ERROR).")
    return p


def main(argv: Optional[Sequence[str]] = None) -> int:
    args = build_arg_parser().parse_args(argv)

    logging.basicConfig(
        level=getattr(logging, args.log_level.upper(), logging.INFO),
        format="%(levelname)s: %(message)s",
    )

    pdf_path = Path(args.pdf).expanduser().resolve()
    if not pdf_path.exists():
        LOGGER.error("PDF not found: %s", pdf_path)
        return 2

    out_dir = Path(args.out).expanduser().resolve()
    ensure_dir(out_dir)

    pages_dir = out_dir / "pages"
    ensure_dir(pages_dir)
    tables_dir = out_dir / "tables"
    images_dir = out_dir / "images"
    renders_dir = out_dir / "renders"

    doc = fitz.open(str(pdf_path))

    doc_meta: Dict[str, Any] = dict(doc.metadata or {})
    doc_meta.update(
        {
            "path": str(pdf_path),
            "page_count": doc.page_count,
        }
    )
    write_json(out_dir / "metadata.json", doc_meta)

    sheets_summary: List[Dict[str, Any]] = []
    sheet_index_pairs_all: List[Tuple[str, str, int]] = []
    images_index: List[Dict[str, Any]] = []

    # Optional JSONL + concatenated text output
    jsonl_path = out_dir / "pages.jsonl"
    jsonl_f = jsonl_path.open("w", encoding="utf-8") if args.jsonl else None
    all_text_chunks: List[str] = []

    page_count = doc.page_count
    max_pages = args.max_pages if args.max_pages and args.max_pages > 0 else page_count

    LOGGER.info("Extracting %s page(s) from %s", min(max_pages, page_count), pdf_path.name)

    for page_idx in range(min(max_pages, page_count)):
        page = doc.load_page(page_idx)
        page_number = page_idx + 1

        full_text = normalize_text(page.get_text("text"))
        if args.ocr_fallback and len(full_text) < args.ocr_min_chars:
            try:
                LOGGER.info("OCR fallback for page %s (text chars=%s)", page_number, len(full_text))
                full_text = ocr_page(page, dpi=args.ocr_dpi)
            except Exception as e:
                LOGGER.warning("OCR failed on page %s: %s", page_number, e)

        page_lines = extract_lines(page) if (args.layout_json or args.sheet_index) else []

        tb = title_block_info(page)

        page_base = pages_dir / f"page_{page_number:03d}"
        write_text(page_base.with_suffix(".txt"), full_text)

        compact = {
            "page_number": page_number,
            "page_index": page_idx,
            "width": float(page.rect.width),
            "height": float(page.rect.height),
            "rotation": int(page.rotation),
            "text_file": page_base.with_suffix(".txt").name,
            "title_block": asdict(tb),
        }
        write_json(page_base.with_suffix(".json"), compact)

        if args.layout_json:
            layout_json = {
                "page_number": page_number,
                "page_index": page_idx,
                "lines": [asdict(l) for l in page_lines],
            }
            write_json(page_base.with_name(page_base.name + "_layout.json"), layout_json)

        if jsonl_f is not None:
            jsonl_obj = dict(compact)
            jsonl_obj["text"] = full_text
            jsonl_f.write(json.dumps(jsonl_obj, ensure_ascii=False) + "\n")

        # Concatenated text (helps quick grepping / indexing)
        header = f"\n\n===== PAGE {page_number:03d} | SHEET {tb.sheet_no or ''} | {tb.sheet_title or ''} =====\n"
        all_text_chunks.append(header + full_text)

        sheet_row = {
            "page_number": page_number,
            "sheet_no": tb.sheet_no or "",
            "sheet_title": tb.sheet_title or "",
            "project_no": tb.project_no or "",
            "project_address": tb.project_address or "",
            "date": tb.date or "",
            "scale": tb.scale or "",
            "titleblock_region": tb.region_name,
        }
        sheets_summary.append(sheet_row)

        if args.sheet_index:
            if "SHEET INDEX" in full_text.upper() or "SHEET LIST" in full_text.upper():
                if not page_lines:
                    page_lines = extract_lines(page)
                pairs = parse_sheet_index_from_page_lines(page_lines)
                for s, t in pairs:
                    sheet_index_pairs_all.append((s, t, page_number))

        if args.images:
            page_img_dir = images_dir / f"page_{page_number:03d}"
            extracted = extract_embedded_images(doc, page, page_img_dir, page_number)
            for rec in extracted:
                rec["sheet_no"] = tb.sheet_no or ""
                rec["sheet_title"] = tb.sheet_title or ""
            images_index.extend(extracted)

        if args.render_pages:
            ensure_dir(renders_dir)
            out_png = renders_dir / f"page_{page_number:03d}.png"
            try:
                render_page(page, out_png, dpi=args.render_dpi)
            except Exception as e:  # pragma: no cover
                LOGGER.warning("Render failed on page %s: %s", page_number, e)

    if jsonl_f is not None:
        jsonl_f.close()

    # Concatenated text file
    write_text(out_dir / "all_text.txt", "".join(all_text_chunks))

    # Sheet summary outputs
    write_json(out_dir / "sheets.json", sheets_summary)
    if sheets_summary:
        with (out_dir / "sheets.csv").open("w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=list(sheets_summary[0].keys()))
            writer.writeheader()
            writer.writerows(sheets_summary)

    sheet_lookup: Dict[int, Dict[str, Any]] = {int(r["page_number"]): r for r in sheets_summary}

    # Sheet index outputs
    sheet_index_records: List[Dict[str, Any]] = [
        {"sheet_no": s, "sheet_title": t, "source_page_number": pnum} for s, t, pnum in sheet_index_pairs_all
    ]
    if sheet_index_records:
        write_json(out_dir / "sheet_index.json", sheet_index_records)
        with (out_dir / "sheet_index.csv").open("w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=list(sheet_index_records[0].keys()))
            writer.writeheader()
            writer.writerows(sheet_index_records)

    # Table extraction
    tables_index: List[Dict[str, Any]] = []
    if args.tables:
        ensure_dir(tables_dir)
        LOGGER.info("Extracting tables with pdfplumber...")
        tables_index.extend(extract_tables_pdfplumber(pdf_path, tables_dir))
        if args.camelot:
            LOGGER.info("Extracting tables with camelot...")
            try:
                tables_index.extend(extract_tables_camelot(pdf_path, tables_dir))
            except Exception as e:
                LOGGER.warning("Camelot table extraction skipped/failed: %s", e)

        # Enrich table records with sheet info
        for rec in tables_index:
            pnum = int(rec.get("page_number", 0))
            info = sheet_lookup.get(pnum)
            if info:
                rec["sheet_no"] = info.get("sheet_no", "")
                rec["sheet_title"] = info.get("sheet_title", "")

        write_json(out_dir / "tables_index.json", tables_index)

    # Images index
    if images_index:
        write_json(out_dir / "images_index.json", images_index)

    # Excel summary
    if args.excel:
        out_xlsx = out_dir / "extracted_summary.xlsx"
        LOGGER.info("Writing Excel summary: %s", out_xlsx.name)
        write_excel_summary(
            out_xlsx=out_xlsx,
            doc_meta=doc_meta,
            sheets=sheets_summary,
            sheet_index_pairs=sheet_index_records,
            tables_index=tables_index,
        )

    LOGGER.info("Done. Output: %s", out_dir)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
